"""
Excel Exporter for IOS Analysis Deliverables.

Generates multi-sheet Excel workbooks with formatted property analysis
per PRD section 9.1 specifications.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# Color definitions for grade-based formatting
GRADE_COLORS = {
    "A": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Green
    "B": PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid"),  # Light green
    "C": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Yellow
    "D": PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid"),  # Light orange
    "F": PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid"),  # Light red
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelExporter:
    """Generate formatted Excel workbooks for IOS property analysis."""

    def __init__(self, output_dir: Path | str = "deliverables"):
        """
        Initialize the Excel exporter.

        Args:
            output_dir: Directory for output files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(
        self,
        scored_df: pd.DataFrame,
        filename: str = "denver_ios_analysis.xlsx",
        filter_criteria: Optional[dict[str, Any]] = None,
    ) -> Path:
        """
        Export scored properties to a multi-sheet Excel workbook.

        Args:
            scored_df: DataFrame with IOS scores and property data
            filename: Output filename
            filter_criteria: Dictionary of filter criteria used (for documentation)

        Returns:
            Path to the generated Excel file
        """
        output_path = self.output_dir / filename
        logger.info(f"Generating Excel workbook: {output_path}")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets in order
        self._create_executive_summary(wb, scored_df, filter_criteria)
        self._create_top_candidates(wb, scored_df)
        self._create_all_properties(wb, scored_df)
        self._create_map_data(wb, scored_df)
        self._create_methodology(wb)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel workbook saved: {output_path}")

        return output_path

    def _create_executive_summary(
        self,
        wb: Workbook,
        df: pd.DataFrame,
        filter_criteria: Optional[dict[str, Any]] = None,
    ) -> None:
        """Create the Executive Summary sheet."""
        ws = wb.create_sheet("Executive Summary")

        # Title
        ws["A1"] = "Denver IOS Property Analysis - Executive Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"].font = Font(italic=True)

        # Overview section
        row = 5
        ws[f"A{row}"] = "ANALYSIS OVERVIEW"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        ws[f"A{row}"] = "Total Properties Analyzed:"
        ws[f"B{row}"] = len(df)
        row += 2

        # Score distribution
        ws[f"A{row}"] = "SCORE DISTRIBUTION"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        grade_col = self._find_column(df, "ios_grade", "grade")
        if grade_col:
            grade_counts = df[grade_col].value_counts().sort_index()
            for grade in ["A", "B", "C", "D", "F"]:
                count = grade_counts.get(grade, 0)
                pct = (count / len(df) * 100) if len(df) > 0 else 0
                ws[f"A{row}"] = f"Grade {grade}:"
                ws[f"B{row}"] = count
                ws[f"C{row}"] = f"{pct:.1f}%"
                if grade in GRADE_COLORS:
                    ws[f"A{row}"].fill = GRADE_COLORS[grade]
                    ws[f"B{row}"].fill = GRADE_COLORS[grade]
                    ws[f"C{row}"].fill = GRADE_COLORS[grade]
                row += 1

        row += 1

        # Property statistics
        ws[f"A{row}"] = "PROPERTY STATISTICS"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # Acreage stats
        acres_col = self._find_column(df, "acres", "calc_acreage", "parcel_acres")
        if acres_col and acres_col in df.columns:
            ws[f"A{row}"] = "Average Parcel Size (acres):"
            ws[f"B{row}"] = round(df[acres_col].mean(), 2)
            row += 1
            ws[f"A{row}"] = "Total Acreage:"
            ws[f"B{row}"] = round(df[acres_col].sum(), 2)
            row += 1

        # Building coverage stats
        coverage_col = self._find_column(df, "building_coverage_pct", "coverage_pct")
        if coverage_col and coverage_col in df.columns:
            ws[f"A{row}"] = "Average Building Coverage:"
            ws[f"B{row}"] = f"{df[coverage_col].mean():.1f}%"
            row += 1

        # Score stats
        score_col = self._find_column(df, "ios_score", "score")
        if score_col and score_col in df.columns:
            ws[f"A{row}"] = "Average IOS Score:"
            ws[f"B{row}"] = round(df[score_col].mean(), 1)
            row += 1
            ws[f"A{row}"] = "Highest IOS Score:"
            ws[f"B{row}"] = round(df[score_col].max(), 1)
            row += 1

        row += 1

        # Filter criteria
        ws[f"A{row}"] = "FILTER CRITERIA USED"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        if filter_criteria:
            for key, value in filter_criteria.items():
                ws[f"A{row}"] = f"{key}:"
                ws[f"B{row}"] = str(value)
                row += 1
        else:
            ws[f"A{row}"] = "No filters applied (all properties included)"
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

    def _create_top_candidates(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create the Top Candidates sheet (A and B grade only)."""
        ws = wb.create_sheet("Top Candidates")

        # Filter for A and B grades
        grade_col = self._find_column(df, "ios_grade", "grade")
        score_col = self._find_column(df, "ios_score", "score")

        if grade_col:
            top_df = df[df[grade_col].isin(["A", "B"])].copy()
        else:
            top_df = df.copy()

        # Sort by score descending
        if score_col:
            top_df = top_df.sort_values(score_col, ascending=False)

        # Add rank column
        top_df = top_df.reset_index(drop=True)
        top_df.insert(0, "Rank", range(1, len(top_df) + 1))

        # Prepare display DataFrame
        display_df = self._prepare_display_dataframe(top_df)

        # Write to sheet
        self._write_dataframe_to_sheet(ws, display_df, apply_grade_colors=True)

        logger.info(f"Top Candidates sheet: {len(top_df)} properties")

    def _create_all_properties(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create the All Properties sheet."""
        ws = wb.create_sheet("All Properties")

        # Sort by score descending
        score_col = self._find_column(df, "ios_score", "score")
        sorted_df = df.copy()
        if score_col:
            sorted_df = sorted_df.sort_values(score_col, ascending=False)

        # Add rank column
        sorted_df = sorted_df.reset_index(drop=True)
        sorted_df.insert(0, "Rank", range(1, len(sorted_df) + 1))

        # Prepare display DataFrame
        display_df = self._prepare_display_dataframe(sorted_df)

        # Write to sheet
        self._write_dataframe_to_sheet(ws, display_df, apply_grade_colors=True)

        logger.info(f"All Properties sheet: {len(sorted_df)} properties")

    def _create_map_data(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create the Map Data sheet (simplified for GIS import)."""
        ws = wb.create_sheet("Map Data")

        # Select and rename columns for GIS
        map_cols = {
            self._find_column(df, "parcel_id", "accountno", "parcelid"): "Parcel_ID",
            self._find_column(df, "address", "situs_address", "full_address"): "Address",
            self._find_column(df, "lat", "latitude", "centroid_lat"): "Latitude",
            self._find_column(df, "lon", "longitude", "lng", "centroid_lon"): "Longitude",
            self._find_column(df, "ios_score", "score"): "IOS_Score",
            self._find_column(df, "ios_grade", "grade"): "Grade",
            self._find_column(df, "acres", "calc_acreage", "parcel_acres"): "Acres",
            self._find_column(df, "building_coverage_pct", "coverage_pct"): "Coverage_Pct",
            self._find_column(df, "zoning_code", "zoning", "zone"): "Zoning",
        }

        # Build map DataFrame
        map_data = {}
        for src_col, dest_col in map_cols.items():
            if src_col and src_col in df.columns:
                map_data[dest_col] = df[src_col]

        map_df = pd.DataFrame(map_data)

        # Write to sheet
        self._write_dataframe_to_sheet(ws, map_df, apply_grade_colors=False)

        logger.info(f"Map Data sheet: {len(map_df)} properties")

    def _create_methodology(self, wb: Workbook) -> None:
        """Create the Methodology sheet with scoring explanation."""
        ws = wb.create_sheet("Methodology")

        methodology_text = """
IOS (Industrial Outdoor Storage) SCORING METHODOLOGY

OVERVIEW
The IOS Score is a composite rating from 0-100 that evaluates each property's
suitability for industrial outdoor storage use. Higher scores indicate better
candidates for IOS investment or development.

SCORING DIMENSIONS (Weighted Components)

1. PARCEL SIZE (25% weight)
   - Optimal: 2-5 acres (score: 100)
   - Excellent: 5-10 acres (score: 95)
   - Good: 1-2 acres (score: 70)
   - Large workable: 10-20 acres (score: 70)
   - Marginal: 0.5-1 acres (score: 30)
   - Too small: <0.5 acres (score: 0)
   - Very large: >20 acres (score: 40)

2. BUILDING COVERAGE (30% weight) - INVERTED
   Lower coverage = higher score (more open space for storage)
   - Optimal: 5-15% coverage (score: 100)
   - Mostly open: <5% (score: 95)
   - Good: 15-25% (score: 85)
   - Marginal: 25-35% (score: 60)
   - Poor: 35-50% (score: 30)
   - Not suitable: >50% (score: 0)

3. ZONING (20% weight)
   - Industrial (I-1, I-2, I-3): score 100
   - Heavy Commercial (C-5): score 75
   - Highway Commercial (C-4): score 65
   - Agricultural (A-1, A-2, A-3): score 55
   - PUD/Mixed Use: score 50
   - Office/Community Commercial: score 40
   - Residential: score 10

4. LAND USE (15% weight)
   Based on current use and property type descriptions.
   - Outdoor storage, contractor yard: score 100
   - Equipment/vehicle storage: score 90-95
   - Industrial yard: score 90
   - Vacant land: score 75
   - Warehouse/distribution: score 55-60
   - Commercial: score 40
   - Residential: score 10

5. STRUCTURAL (5% weight)
   Based on building count and sizes.
   - Vacant (no buildings): +20 bonus
   - Single small building (<2,000 sqft): +30 total bonus
   - Multiple or large buildings: penalties applied

6. LOCATION (5% weight)
   - Base score for target industrial area: 60
   - DIA proximity bonus: +15

GRADE CLASSIFICATION

- Grade A (85-100): Excellent IOS Candidate
  High priority - immediate follow-up recommended

- Grade B (75-84): Good IOS Candidate
  Strong potential - worth detailed analysis

- Grade C (65-74): Moderate IOS Candidate
  Possible opportunity - review for specific use cases

- Grade D (50-64): Marginal IOS Candidate
  Limited suitability - consider only if other factors favorable

- Grade F (0-49): Poor IOS Candidate
  Not recommended for IOS use

DATA SOURCES

- Parcel boundaries and ownership: Adams County Assessor
- Building footprints: Adams County GIS
- Zoning: Adams County Planning & Development
- Sales history: Adams County Assessor (when available)

NOTES

- Scores are relative rankings, not absolute valuations
- Manual verification recommended for top candidates
- Zoning changes or variances may affect actual suitability
- Building coverage calculated from GIS building footprints

Generated by Denver IOS Property Sourcing System
"""

        lines = methodology_text.strip().split("\n")
        for i, line in enumerate(lines, start=1):
            ws[f"A{i}"] = line
            if line.isupper() and line.strip():
                ws[f"A{i}"].font = Font(bold=True)

        ws.column_dimensions["A"].width = 80

    def _prepare_display_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepare a DataFrame for display with ALL columns, priority columns first."""

        # Priority column mappings (source options -> display name)
        # These appear first in the output, in this order
        priority_mappings = [
            # Core
            (["Rank"], "Rank"),
            (["parcel_id", "PARCELNB"], "Parcel ID"),
            (["PIN"], "PIN"),

            # IOS Scoring
            (["ios_score", "score"], "IOS Score"),
            (["ios_grade", "grade"], "Grade"),
            (["ios_tier"], "Tier"),
            (["ios_notes", "scoring_notes"], "Analysis Notes"),
            (["score_parcel_size"], "Score: Parcel Size"),
            (["score_building_coverage"], "Score: Bldg Coverage"),
            (["score_zoning"], "Score: Zoning"),
            (["score_land_use"], "Score: Land Use"),
            (["score_structural"], "Score: Structural"),
            (["score_location"], "Score: Location"),

            # Address
            (["concataddr1"], "Address"),
            (["concataddr2"], "Address Line 2"),
            (["streetno"], "Street Number"),
            (["streetdir"], "Street Direction"),
            (["streetname"], "Street Name"),
            (["streetsuf"], "Street Suffix"),
            (["streetpostdir"], "Street Post Dir"),
            (["streetalp"], "Street Alpha"),
            (["loccity"], "City"),
            (["loczip"], "Zip"),

            # Parcel Size & Coverage
            (["parcel_area_sqft"], "Parcel Sqft"),
            (["lot_size"], "Lot Size"),
            (["lot_measure_unit"], "Lot Unit"),
            (["building_coverage_pct", "coverage_pct"], "Building Coverage %"),

            # Building Info
            (["building_footprint_count"], "Building Count"),
            (["building_footprint_sqft"], "Building Footprint Sqft"),
            (["total_building_sqft"], "Total Building Sqft"),
            (["improvement_count"], "Improvement Count"),
            (["building_description"], "Building Description"),
            (["oldest_year_built"], "Year Built"),
            (["bedrooms"], "Bedrooms"),
            (["bathrooms"], "Bathrooms"),
            (["rooms"], "Rooms"),
            (["vacant_improved"], "Vacant/Improved"),

            # Zoning
            (["zoning_code", "zoning", "zone"], "Zoning Code"),
            (["zoning_jurisdiction"], "Zoning Jurisdiction"),

            # Property Type
            (["property_type", "accttype"], "Property Type"),

            # Values
            (["actual_total_value"], "Actual Total Value"),
            (["actual_land_value"], "Actual Land Value"),
            (["actual_improvement_value"], "Actual Improvement Value"),
            (["assessed_total_value"], "Assessed Total Value"),

            # Sales
            (["last_sale_date"], "Last Sale Date"),
            (["last_sale_price"], "Last Sale Price"),
            (["last_deed_type"], "Last Deed Type"),
            (["buyer"], "Buyer"),
            (["seller"], "Seller"),

            # Owner Info
            (["ownernamefull"], "Owner Name"),
            (["ownername1"], "Owner Name 1"),
            (["ownername2"], "Owner Name 2"),
            (["owneraddressfull"], "Owner Address Full"),
            (["owneraddress"], "Owner Address"),
            (["ownercity"], "Owner City"),
            (["ownerstate"], "Owner State"),
            (["ownerzip"], "Owner Zip"),
            (["ownerpostalcode"], "Owner Postal Code"),
            (["ownercpp"], "Owner CPP"),
            (["ownercsz"], "Owner City/State/Zip"),
            (["ownerprovince"], "Owner Province"),
            (["ownercountry"], "Owner Country"),

            # Legal
            (["legal"], "Legal Description"),
            (["subname"], "Subdivision Name"),

            # Shape
            (["Shape_Area"], "Shape Area"),
            (["Shape_Length"], "Shape Length"),
        ]

        result_data = {}
        used_source_cols = set()

        # First pass: add priority columns in order
        for source_cols, display_name in priority_mappings:
            found_col = None
            for col in source_cols:
                if col in df.columns:
                    found_col = col
                    break

            if found_col:
                result_data[display_name] = df[found_col]
                used_source_cols.add(found_col)

        # Computed columns

        # Calculate acres from sqft
        parcel_sqft_col = self._find_column(df, "parcel_area_sqft")
        if parcel_sqft_col and parcel_sqft_col in df.columns:
            result_data["Acres"] = df[parcel_sqft_col] / 43560.0

        # Calculate open space %
        coverage_col = self._find_column(df, "building_coverage_pct", "coverage_pct")
        if coverage_col and coverage_col in df.columns:
            result_data["Open Space %"] = 100 - df[coverage_col]

        # Generate coordinates from geometry
        if "geometry" in df.columns:
            try:
                centroids = df.geometry.centroid
                result_data["Lat"] = centroids.y
                result_data["Lon"] = centroids.x
                used_source_cols.add("geometry")

                # Google Maps links
                result_data["Google Maps Link"] = [
                    f"https://www.google.com/maps?q={lat},{lon}"
                    if pd.notna(lat) and pd.notna(lon) else ""
                    for lat, lon in zip(centroids.y, centroids.x)
                ]
            except Exception:
                pass

        # Calculate price per acre
        land_col = self._find_column(df, "actual_land_value")
        if land_col and land_col in df.columns and "Acres" in result_data:
            result_data["Price Per Acre"] = [
                round(land / acres, 0) if pd.notna(land) and pd.notna(acres) and acres > 0 else None
                for land, acres in zip(df[land_col], result_data["Acres"])
            ]

        # Add any remaining columns not yet included
        for col in df.columns:
            if col not in used_source_cols and col != "geometry":
                # Check if this column was already mapped
                already_added = False
                for source_cols, _ in priority_mappings:
                    if col in source_cols:
                        already_added = True
                        break

                if not already_added:
                    # Add with a clean display name
                    display_name = col.replace("_", " ").title()
                    if display_name not in result_data:
                        result_data[display_name] = df[col]

        return pd.DataFrame(result_data)

    def _write_dataframe_to_sheet(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        apply_grade_colors: bool = True,
    ) -> None:
        """Write a DataFrame to a worksheet with formatting."""
        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Write data rows
        grade_col_idx = None
        if "Grade" in df.columns:
            grade_col_idx = list(df.columns).index("Grade")

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            grade = None
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Handle different value types
                if pd.isna(value):
                    cell.value = ""
                elif isinstance(value, float):
                    # Format based on column name
                    col_name = df.columns[col_idx - 1]
                    if "%" in col_name or col_name in ["Building Coverage %", "Open Space %"]:
                        cell.value = round(value, 1)
                        cell.number_format = "0.0%"
                    elif col_name in ["Total Assessed Value", "Land Value", "Price Per Acre",
                                      "Actual Total Value", "Actual Land Value",
                                      "Actual Improvement Value", "Assessed Total Value",
                                      "Last Sale Price"]:
                        cell.value = round(value, 0)
                        cell.number_format = "$#,##0"
                    elif col_name in ["Acres", "IOS Score"]:
                        cell.value = round(value, 2)
                    else:
                        cell.value = value
                elif hasattr(value, 'tzinfo') and value.tzinfo is not None:
                    # Handle timezone-aware datetime by removing timezone
                    cell.value = value.replace(tzinfo=None)
                elif isinstance(value, (pd.Timestamp,)):
                    # Handle pandas Timestamp
                    if value.tzinfo is not None:
                        cell.value = value.tz_localize(None).to_pydatetime()
                    else:
                        cell.value = value.to_pydatetime()
                else:
                    cell.value = value

                # Track grade for row coloring
                if col_idx - 1 == grade_col_idx:
                    grade = value

                cell.border = THIN_BORDER

            # Apply grade-based row coloring
            if apply_grade_colors and grade in GRADE_COLORS:
                for col_idx in range(1, len(row) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = GRADE_COLORS[grade]

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_length = len(str(col_name))
            for row_idx in range(2, min(102, len(df) + 2)):  # Check first 100 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max_length + 2, 50
            )

        # Freeze header row
        ws.freeze_panes = "A2"

    def _find_column(self, df: pd.DataFrame, *options: str) -> Optional[str]:
        """Find the first matching column from a list of options."""
        for opt in options:
            if opt in df.columns:
                return opt
            # Try case-insensitive match
            for col in df.columns:
                if col.lower() == opt.lower():
                    return col
        return None


def export_to_excel(
    scored_df: pd.DataFrame,
    output_dir: Path | str = "deliverables",
    filename: str = "denver_ios_analysis.xlsx",
    filter_criteria: Optional[dict[str, Any]] = None,
) -> Path:
    """
    Convenience function to export scored properties to Excel.

    Args:
        scored_df: DataFrame with IOS scores and property data
        output_dir: Output directory
        filename: Output filename
        filter_criteria: Filter criteria used (for documentation)

    Returns:
        Path to generated Excel file
    """
    exporter = ExcelExporter(output_dir)
    return exporter.export(scored_df, filename, filter_criteria)
